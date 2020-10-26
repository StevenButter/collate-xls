from openpyxl import load_workbook


def GetWorkbookAsDictionary(filename):
    customerWorkbook = load_workbook(filename)
    customerSheet = customerWorkbook.active

    return [_GetDictForRow(row, customerSheet)
            for row in customerSheet.iter_rows(min_row=2)]


def _GetDictForRow(row, sheet):
    return {sheet.cell(1, cell.column).value: cell.value for cell in row}
