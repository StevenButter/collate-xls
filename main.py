from openpyxl import load_workbook

# column 10 is missing in customer_copy


def GetDictForRow(row, sheet):
    return {customerSheet.cell(1, cell.column).value: cell.value for cell in row}


if __name__ == "__main__":
    customerWorkbook = load_workbook('customer_copy.xlsx')
    customerSheet = customerWorkbook.active

    rows = [GetDictForRow(row, customerSheet)
            for row in customerSheet.iter_rows(min_row=2)]